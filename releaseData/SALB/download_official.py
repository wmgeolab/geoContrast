'''
Downloads the latest available shapefiles/zipfiles for each country
and automatically generates the sourceMetaData.json file.

NOTE: the dates on the website and filenames are the start dates of temporal validity, not the end dates.
Metadata for year and source_updated has to be manually input by looking at the historical change table
on the country pages.
'''


import urllib.request
from zipfile import ZipFile
import os
import io
import sys
import json
import logging
import datetime

# import iotools.py
# should be in the top repo folder
cur_dir = os.path.abspath('')
repo_dir = cur_dir.split('sourceData')[0]
sys.path.append(repo_dir)
import iotools

# redirect to logfile
logger = open('download_log.txt', mode='w', encoding='utf8', buffering=1)
sys.stdout = logger
sys.stderr = logger

# access the gadm country download page
root = 'https://www.unsalb.org'

def parse_country_links(raw):
    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # get all country links
    for elem in elems:
        if elem.startswith('a href="/data'):
            relUrl = elem.replace('a href=', '').strip('"')
            url = root + relUrl
            yield url

def iter_country_page_downloads(url):
    raw = urllib.request.urlopen(url).read().decode('utf8')

    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # find all zipfile downloads (should contain shapefile)
    for elem in elems:
        if elem.startswith('span class="date-display'):
            fromdate,todate = next(elems).split(' to ')
            # fromdate
            yr,mn,dy = fromdate.split('/')
            yr,mn,dy = map(int, [yr,mn,dy])
            fromdate = datetime.date(yr, mn, dy)
            # todate
            todate = todate if todate != 'last update' else None
            if todate:
                yr,mn,dy = todate.split('/')
                yr,mn,dy = map(int, [yr,mn,dy])
                todate = datetime.date(yr, mn, dy)
            print('FROM DATE',fromdate,'TO DATE',todate)
        if elem.startswith("a class='file'"):
            url = elem.replace("a class='file' href=", "").strip("'") # the url tag oddly uses single-quotes
            print('FILE',url)
            if url.endswith('.zip'):
                yield fromdate,todate, url

def get_historical_table_download(url):
    raw = urllib.request.urlopen(url).read().decode('utf8')

    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # find all excel downloads
    downloads = []
    for elem in elems:
        elem = elem.replace('"','').replace("'","")
        if elem.endswith('.xlsx'):
            _url = elem.split('href=')[-1]
            downloads.append(_url)
    if len(downloads) > 1:
        logging.warning('More than one historical tables, use only the first one')

    # return first download
    return downloads[0] if downloads else None

def parse_historical_table(url):
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(urllib.request.urlopen(url).read()))
    # get metadata sheet
    metasheet = None
    for name in wb.sheetnames:
        if name.lower().startswith('metadata'):
            metasheet = wb[name]
            break
    # otherwise return early
    if not metasheet:
        logging.warning('Historal table had no metadata sheet')
        return {}
    # extract values from key-value row entries
    info = {}
    for row in metasheet.iter_rows():
        for i,cell in enumerate(row):
            if cell.value == 'Last update':
                updated = row[i+1].value
                info['updated'] = updated
    return info
        

# loop pages and download+unzip each
print('downloading:')
page = 0 # starts at page 0
while True:
    url = '{}/data?page={}'.format(root, page)
    print('')
    print('looping country links from page', url)
    resp = urllib.request.urlopen(url)
    if resp.getcode() != 200:
        # reached the end/invalid page
        break

    # get country links
    raw = resp.read().decode('utf8')
    countrylinks = list(parse_country_links(raw))
    if len(countrylinks) == 0:
        # reached the end/invalid page
        break
    
    # loop
    for countrylink in countrylinks:
        print(countrylink)
        
        # get page downloads
        page_downloads = list(iter_country_page_downloads(countrylink))
        if len(page_downloads) == 0:
            print('No page downloads, skipping')
            continue
        if len(page_downloads) > 1:
            logging.warning('More than one country page downloads: {}'.format(page_downloads))
        fromdate,todate,ziplink = page_downloads[0] # first one should be the most recent
        
        # create country folder
        iso = countrylink[-3:].upper()
        if os.path.lexists(iso):
            # dont overwrite any existing folders
            # to protect any metadata that may have been manually edited
            print('Country folder already exists, skipping')
            continue
        else:
            os.mkdir(iso)
        
        # download zipfile
        zipname = ziplink.split('/')[-1]
        urllib.request.urlretrieve(ziplink, '{}/{}'.format(iso, zipname))
        
        # determine main entries for meta file
        archive = ZipFile(os.path.join(iso, zipname), 'r')
        shapefiles = [name
                     for name in archive.namelist()
                     if name.endswith('.shp')]
        if len(shapefiles) == 0:
            logging.warning('Zipfile does not contain any shapefile: {}'.format(archive.namelist()))
            continue
        if len(shapefiles) > 1:
            logging.warning('Zipfile contains more than one shapefile: {}'.format(shapefiles))
            # in most cases this appears to be two files 'BNDA_*' and 'BNDL_*'
            # it appears the one we want is 'BNDA_*'
            # sorting should fix it
            shapefile_name = sorted(shapefiles)[0]
        else:
            shapefile_name = shapefiles[0]
        shapefile_path = '{}/{}'.format(zipname, shapefile_name)

        # read shapefile info
        reader = iotools.get_reader(os.path.join(iso, shapefile_path))
        fieldnames = [f[0] for f in reader.fields]

        # if present, use download link's "to" date text as basis for source_date and year
        # in most cases the todate is set to "last update"
        # but in cases where the shapefile has not been uploaded yet
        # we can use the todate of a previously uploaded shapefile
        updated = None
        year = None
        if todate:
            year = todate.year
            updated = todate.isoformat()
            print('''source_update and year set to download link's "to" date text''')

        # compare with date info from historical excel table
        table_download = get_historical_table_download(countrylink)
        table_updated = None
        if table_download:
            table_info = parse_historical_table(table_download)
            table_updated = table_info.get('updated', None)
            if table_updated:
                print('detected historical excel table "Last updated" field: {}'.format(table_updated))
        else:
            logging.warning('could not find historical table download')

        # fallback on historical table field if date/year hasn't been found yet
        if not updated and table_updated:
            if hasattr(table_updated, 'year'):
                # already date or datetime value
                yr,mn,dy = table_updated.year, table_updated.month, table_updated.day
                updated = datetime.date(yr, mn, dy).isoformat()
            else:
                # date string
                yr,mn,dy = table_updated.split('-')
                dy,mn,yr = map(int, [dy,mn,yr])
                updated = datetime.date(yr, mn, dy).isoformat()
            year = yr
            print('source_update and year set to historical excel table "Last updated" field')

        # compare with "DATSOR" field in shapefile
        datsor = None
        if 'DATSOR' in fieldnames:
            for rec in reader.iterRecords():
                datsor = rec['DATSOR']
                if datsor and len(datsor.split('/')) == 3:
                    break
            if datsor:
                print('detected "DATSOR" field in shapefile: {}'.format(datsor))
            
        # fallback on datsor field if date/year hasn't been found yet
        if not updated and datsor:
            dy,mn,yr = datsor.split('/')
            dy,mn,yr = map(int, [dy,mn,yr])
            updated = datetime.date(yr, mn, dy).isoformat()
            year = yr
            print('source_update and year set to shapefile "DATSOR" field')

        # if nothing else, fallback on download link's "from" date text
        if not updated:
            year = fromdate.year
            updated = fromdate.isoformat()
            logging.warning('''source_update and year set to download link's "from" date text''')

        # get admin level
        # all salb files have both ADM1NM and ADM2NM
        # in rare cases, ADM2NM will be empty, indicating an adm1 boundary
        rec = reader.record(0)
        if rec['ADM2NM']:
            level = 2
        else:
            level = 1
        reader.close()

        # create metadata
        meta = {
            "input":[
                    {'path':shapefile_path,
                    'level':lev,
                    "name_field":'ADM{}NM'.format(lev) if lev > 0 else None,
                    "dissolve":'ADM{}NM'.format(lev) if lev > 0 else True,
                    "keep_fields":[f for f in fieldnames if f=='ISO3CD' or (f.startswith('ADM') and int(f[3]) <= lev) ],
                    }
                    for lev in range(level+1)
                ],
            
            "iso":iso,
            "year":year,
            "type_field":None,
            
            "source":["UN SALB"],
            "source_updated":updated,
            "source_url":countrylink,
            "download_url":countrylink,
            "license":'UN SALB Data License',
            "license_detail":None,
            "license_url":"https://www.unsalb.org/sites/default/files/wysiwyg_uploads/docs_uploads/TermsOfUseSALB2021.pdf"
        }
        print(meta)

        # write metadata to file
        dst = os.path.join(iso, 'sourceMetaData.json')
        with open(dst, 'w', encoding='utf8') as fobj:
            json.dump(meta, fobj, indent=4)

    page += 1
        
